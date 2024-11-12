# -*- coding: utf-8 -*-
from functools import reduce
import os
import holidays
import numpy as np
import pandas as pd
import dill
import openpyxl

# Set the base directory to the 'frm/utils' directory relative to the current module location
utils_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '.'))
file_path_locale_holiday = os.path.join(utils_dir, 'LOCALE_HOLIDAY.pkl')
file_path_ccy_holiday = os.path.join(utils_dir, 'CCY_HOLIDAY.pkl')

with open(file_path_locale_holiday, 'rb') as f:
    LOCALE_HOLIDAY = dill.load(f)

with open(file_path_ccy_holiday, 'rb') as f:
    CCY_HOLIDAY = dill.load(f)


def get_holidays_object(key):
    if key in CCY_HOLIDAY.keys():
        return CCY_HOLIDAY[key]
    elif key in LOCALE_HOLIDAY.keys():
        return LOCALE_HOLIDAY[key]
    else:
        raise ValueError(f"Holidays not setup for {key}")


def convert_weekend_to_weekmask(weekend_set):
    "For converting the holidays.weekend attribute to a valid weekmask for numpy.busdaycalendar"
    # Initialize a list of ones representing Monday to Sunday (working days)
    weekmask = [1, 1, 1, 1, 1, 1, 1]
    
    # Loop through the weekend days and set corresponding indexes in the weekmask to 0
    for weekend_day in weekend_set:
        weekmask[weekend_day] = 0
    
    # Return the weekmask as a string (optional, can also return as a list if needed)
    return ''.join(map(str, weekmask))


def get_busdaycal(keys) -> np.busdaycalendar:
    """
    Create a calendar which has the holidays and business days of the currencies/locales.
    """

    if keys is None:
        return np.busdaycalendar()

    if type(keys) is str:
        keys = [keys.upper()]
    elif type(keys) is list:
        keys = list(set(keys))
        keys = [key.upper() for key in keys]

    holiday_objects = [get_holidays_object(key) for key in keys]
    combined_weekend_union = reduce(lambda x, y: x | y.weekend, holiday_objects, set())
    weekmasks = convert_weekend_to_weekmask(combined_weekend_union)

    # Flattan the list of lists
    holidays = [h for holiday_list in holiday_objects for h in holiday_list]

    return np.busdaycalendar(weekmask=weekmasks, holidays=holidays)  



# Pickle variables 
if __name__ == "__main__":
    
    years = range(1990,2100)

    LOCALE_HOLIDAY = dict(sorted({
        'AE-DUBAI': holidays.AE(categories=['public'], years=years),
        'AR-BUENOS_AIRES': holidays.AR(categories=['public'], years=years),
        'AU-SYDNEY': holidays.AU(subdiv='NSW', categories=['public', 'bank'], years=years),
        'AU-MELBOURNE': holidays.AU(subdiv='VIC', categories=['public', 'bank'], years=years),
        'BR-SAO_PAULO': holidays.BR(subdiv='SP', categories=['public'], years=years),
        'CA-TORONTO': holidays.CA(subdiv='ON', categories=['public'], years=years),
        'CH-ZURICH': holidays.CH(subdiv='ZH', categories=['public'], years=years),
        'CL-SANTIAGO': holidays.CL(subdiv='RM', categories=['public', 'bank'], years=years),
        'CN-SHANGHAI': holidays.CN(categories=['public'], years=years),
        'CO-BOGOTA': holidays.CO(categories=['public'], years=years),
        'CZ-PRAGUE': holidays.CZ(categories=['public'], years=years),
        'DK-COPENHAGEN': holidays.DK(categories=['public'], years=years),
        'DE-FRANKFURT': holidays.DE(subdiv='HH', categories=['public'], years=years),
        'ES-MADRID': holidays.ES(subdiv='MD', categories=['public'], years=years),
        'FR-PARIS': holidays.FR(categories=['public'], years=years),
        'FI-HELSINKI': holidays.FI(categories=['public'], years=years),
        'HK-HONG_KONG': holidays.HK(categories=['public'], years=years),
        'HU-BUDAPEST': holidays.HU(categories=['public'], years=years),
        'ID-JAKARTA': holidays.ID(categories=['public'], years=years),
        'IN-MUMBAI': holidays.IN(subdiv='MH', categories=['public'], years=years),
        'IL-TEL_AVIV': holidays.IL(categories=['public'], years=years),
        'IS-REYKJAVIK': holidays.IS(categories=['public'], years=years),
        'JP-TOYKO': holidays.JP(categories=['public'], years=years),
        'MY-KUALA_LUMPUR': holidays.MY(subdiv='KUL', categories=['public'], years=years),
        'MX-MEXICO_CITY': holidays.MX(categories=['public'], years=years),
        'NO-OSLO': holidays.NO(categories=['public'], years=years),
        'NZ-AUCKLAND': holidays.NZ(subdiv='AUK', categories=['public'], years=years),
        'PL-WARSAW': holidays.PL(categories=['public'], years=years),
        'PH-MANILA': holidays.PH(categories=['public'], years=years),
        'RU-MOSCOW': holidays.RU(categories=['public'], years=years),
        'SA-RIYADH': holidays.SA(categories=['public'], years=years),
        'SE-STOCKHOLM': holidays.SE(categories=['public'], years=years),
        'SG-SINGAPORE': holidays.SG(categories=['public'], years=years),
        'TH-BANKCOK': holidays.TH(categories=['public', 'bank'], years=years),
        'TR-ISTANBUL': holidays.TR(categories=['public'], years=years),
        'TW-TAPEI': holidays.TW(categories=['public'], years=years),
        'UK-LONDON': holidays.UK(subdiv='ENG', categories=['public'], years=years),
        'US-NEW_YORK': holidays.US(subdiv='NY', categories=['public'], years=years),
        'VN-HO_CHI_MINH_CITY': holidays.VN(categories=['public'], years=years),
        'ZA-JOHANNESBURG': holidays.ZA(categories=['public'], years=years),
        'European_Central_Bank': holidays.ECB(categories=['public'], years=years),
        'ICE_Futures_Europe': holidays.IFEU(categories=['public'], years=years), # Requires holidays >=0.49
        'New_York_Stock_Exchange': holidays.XNYS(categories=['public'], years=years),
    }.items()))

    CCY_HOLIDAY = dict(sorted({
        'AED': holidays.AE(categories=['public'], years=years),
        'ARS': holidays.AR(categories=['public'], years=years),
        'AUD': holidays.AU(subdiv='NSW', categories=['public', 'bank'], years=years),
        'BRL': holidays.BR(subdiv='SP', categories=['public'], years=years),
        'CAD': holidays.CA(subdiv='ON', categories=['public'], years=years),
        'CHF': holidays.CH(subdiv='ZH', categories=['public'], years=years),
        'CLP': holidays.CL(subdiv='RM', categories=['public', 'bank'], years=years),
        'CNY': holidays.CN(categories=['public'], years=years),
        'COP': holidays.CO(categories=['public'], years=years),
        'CZK': holidays.CZ(categories=['public'], years=years),
        'DKK': holidays.DK(categories=['public'], years=years),
        'HKD': holidays.HK(categories=['public'], years=years),
        'HUF': holidays.HU(categories=['public'], years=years),
        'IDR': holidays.ID(categories=['public'], years=years),
        'INR': holidays.IN(subdiv='MH', categories=['public'], years=years),
        'ILS': holidays.IL(categories=['public'], years=years),
        'ISK': holidays.IS(categories=['public'], years=years),
        'JPY': holidays.JP(categories=['public'], years=years),
        'MYR': holidays.MY(subdiv='KUL', categories=['public'], years=years),
        'MXN': holidays.MX(categories=['public'], years=years),
        'NOK': holidays.NO(categories=['public'], years=years),
        'NZD': holidays.NZ(subdiv='AUK', categories=['public'], years=years),
        'PLN': holidays.PL(categories=['public'], years=years),
        'PHP': holidays.PH(categories=['public'], years=years),
        'RUB': holidays.RU(categories=['public'], years=years),
        'SAR': holidays.SA(categories=['public'], years=years),
        'SEK': holidays.SE(categories=['public'], years=years),
        'SGD': holidays.SG(categories=['public'], years=years),
        'THB': holidays.TH(categories=['public', 'bank'], years=years),
        'TRY': holidays.TR(categories=['public'], years=years),
        'TWD': holidays.TW(categories=['public'], years=years),
        'GBP': holidays.UK(subdiv='ENG', categories=['public'], years=years),
        'USD': holidays.US(subdiv='NY', categories=['public'], years=years),
        'VND': holidays.VN(categories=['public'], years=years),
        'ZAR': holidays.ZA(categories=['public'], years=years),
        'EUR': holidays.ECB(categories=['public'], years=years),
    }.items()))

    # Pickle the variable
    with open(file_path_locale_holiday, 'wb') as f:
        dill.dump(LOCALE_HOLIDAY, f)

    # Pickle the variable
    with open(file_path_ccy_holiday, 'wb') as f:
        dill.dump(CCY_HOLIDAY, f)
        

if __name__ == "__main__":

    # Create code for static holidays definition
    df = pd.read_excel(utils_dir + "\\calendar_map.xlsx", usecols=range(6))
    holiday_dict = {}
    holiday_dict_str = "LOCALE_HOLIDAY = dict(sorted({\n"
    ccy_holiday_dict = {}
    ccy_holiday_dict_str = "CCY_HOLIDAY = dict(sorted({\n"

    for i, row in df.iterrows():
        eval_str = None
        ccy = row["ISOCurrencyCode"]
        ui_string = row["UIString"]
        code = row["Code"]
        subdivision = row["Subdivision"]

        try:
            categories = [holidays.PUBLIC, holidays.BANK]
            if pd.isna(row["Subdivision"]):
                eval_str = f"holidays.{code}(categories={categories}, years=years)"

            else:
                eval_str = (
                    f"holidays.{code}(subdiv='{subdivision}', categories={categories}, years=years)"
                )
            holiday_obj = eval(eval_str)

        except ValueError:
            categories = [holidays.PUBLIC]
            if pd.isna(row["Subdivision"]):
                eval_str = f"holidays.{code}(categories={categories}, years=years)"
            else:
                eval_str = (
                    f"holidays.{code}(subdiv='{subdivision}', categories={categories}, years=years)"
                )
            holiday_obj = eval(eval_str)

        holiday_dict[ui_string] = holiday_obj
        holiday_dict_str += f"    '{ui_string}': {eval_str},\n"

        if row["CurrencyPrimaryCalendar"]:
            ccy_holiday_dict[ccy] = holiday_obj
            ccy_holiday_dict_str += f"    '{ccy}': {eval_str},\n"

    holiday_dict_str += "}.items()))"
    ccy_holiday_dict_str += "}.items()))"

    print(holiday_dict_str)
    print("\n")
    print(ccy_holiday_dict_str)




    # Initialize Excel writer
    # Prepare the output file path
    output_path = os.path.join(utils_dir, 'currency_holidays.xlsx')

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'holidays'

    # Define column mappings for each currency
    col_start = 1  # Initial column start in openpyxl (1-based index)

    for ccy in ['aud', 'eur', 'gbp', 'jpy', 'nzd', 'usd']:
        # Prepare data for each currency
        dates = list(CCY_HOLIDAY[ccy.upper()].keys())
        holiday_name = list(CCY_HOLIDAY[ccy.upper()].values())

        # Create a DataFrame with currency-specific columns
        holiday_df = pd.DataFrame({
            f'{ccy.upper()}_date': dates,
            f'{ccy.upper()}_holiday_name': holiday_name
        })

        # Sort the DataFrame by date
        holiday_df.sort_values(by=f'{ccy.upper()}_date', inplace=True)
        holiday_df.reset_index(drop=True, inplace=True)

        # Write headers
        ws.cell(row=1, column=col_start, value=f'{ccy.upper()}_date')
        ws.cell(row=1, column=col_start + 1, value=f'{ccy.upper()}_holiday_name')

        # Write sorted data
        for i, row in holiday_df.iterrows():
            ws.cell(row=i + 2, column=col_start, value=row[f'{ccy.upper()}_date'])
            ws.cell(row=i + 2, column=col_start + 1, value=row[f'{ccy.upper()}_holiday_name'])

        # Move to the next pair of columns for the next currency
        col_start += 2

    # Save the workbook
    wb.save(output_path)

    # Define the output path
    #output_path = os.path.join(utils_dir, 'currency_holidays.xlsx')

    # Write to an Excel file, starting from column A
    #combined_df.to_excel(output_path, index=False, sheet_name='Currency Holidays')